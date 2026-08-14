"""Matriz de compatibilidade padrão de entrada × cargas → inversores que atendem.

Formato de validação pedido pela engenharia: para cada combinação de rede
(tensão + tipo de ligação) com cada tipo de carga de backup (tensão + fase),
listar TODOS os inversores híbridos e TODOS os inversores on-grid que atendem.

O ponto da tabela é auditar o motor, então ela **chama as funções do motor** —
`_bloqueio_rede`, `compativel_com_cargas` e o mapa `CONEXOES_REDE`. Nada da
lógica de compatibilidade é reimplementado aqui; se a regra mudar no motor, a
matriz muda junto. Uma matriz que reimplementasse as regras não validaria nada:
concordaria consigo mesma.

Uso:
    DATABASE_URL=... python scripts/matriz_compatibilidade.py [--xlsx CAMINHO]

Sem --xlsx, imprime no terminal.
"""

import asyncio
import os
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

if sys.platform == "win32":
    asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

from app.calculate.service import CONEXOES_REDE                     # noqa: E402
from app.catalog.service import list_products                       # noqa: E402
from app.database import AsyncSessionLocal                          # noqa: E402
from app.engines.kit_attributes import eff, preco_venda              # noqa: E402
from app.engines.kit_builder import (                               # noqa: E402
    FASES_REDE,
    _NOME_FASES,
    _alertas_rede,
    _bloqueio_rede,
    carga_existe_na_rede,
    compativel_com_cargas,
)

# ── eixos da matriz ──────────────────────────────────────────────────────────

#: (chave interna, tensão exibida, ligação exibida)
PADROES = [
    ("mono_127",    "127",     "Monofásico"),
    ("mono_220",    "220",     "Monofásico"),
    ("bi_127_220",  "127/220", "Bifásico"),
    ("bi_220_380",  "220/380", "Bifásico"),
    ("tri_127_220", "127/220", "Trifásico"),
    ("tri_220_380", "220/380", "Trifásico"),
]

#: (tensão da carga, fase da carga, rótulo exibido). None = sem cargas de backup.
CARGAS = [
    (None,  None,         "— (sem cargas de backup)"),
    ("127", "monofasico", "Monofásico"),
    ("220", "monofasico", "Monofásico"),
    ("220", "bifasico",   "Bifásico"),
    ("380", "bifasico",   "Bifásico"),
    ("220", "trifasico",  "Trifásico"),
    ("380", "trifasico",  "Trifásico"),
]


def _tit(p) -> str:
    return str(eff(p, "title") or getattr(p, "meubess_id", "?"))


def _hibridos_que_atendem(inversores, padrao, fase_rede, tensao_carga, fase_carga):
    """Quais híbridos servem essa combinação, e por que os outros não.

    Reproduz a ordem de filtros do build_kits: primeiro a rede (R7), depois as
    cargas (R8). Potência fica de fora de propósito — a matriz é de
    compatibilidade elétrica, não de dimensionamento; qualquer modelo pode ser
    escalado em paralelo ou combinado com mais baterias.

    Devolve três listas, não duas. `com_ressalva` são os que passam nos filtros
    mas disparam alerta de rede — hoje o caso do inversor trifásico em unidade
    monofásica, que o motor NÃO bloqueia (exige mudança do padrão de entrada
    junto à concessionária). Misturá-los com os que atendem direto faria a
    matriz afirmar algo que a instalação não sustenta.
    """
    atendem, com_ressalva, recusados = [], [], []
    tensoes = {tensao_carga} if tensao_carga else None
    fases = {fase_carga} if fase_carga else None
    # Só carga trifásica exige casar a tensão ENTRE FASES (ver _tensoes_entre_fases).
    entre_fases = {tensao_carga} if (tensao_carga and fase_carga == "trifasico") else None

    for inv in inversores:
        motivo = _bloqueio_rede(inv, padrao)
        if motivo:
            recusados.append((_tit(inv), motivo))
            continue
        if not eff(inv, "eps_output_voltage"):
            recusados.append((_tit(inv), "tensão de saída EPS não cadastrada"))
            continue
        # padrao entra aqui porque a saída EPS das linhas K é selecionável e a
        # seleção já foi consumida pela ligação na rede (ver _eps_efetivo).
        ok, why = compativel_com_cargas(inv, tensoes, fases, entre_fases, padrao)
        if not ok:
            recusados.append((_tit(inv), why or "incompatível"))
            continue
        ressalvas = [a for a in _alertas_rede(inv, fase_rede, padrao)
                     if "requer aumento de carga" in a]
        (com_ressalva if ressalvas else atendem).append(_tit(inv))
    return atendem, com_ressalva, recusados


def _ongrid_que_atendem(inversores, padrao):
    """Quais inversores string se conectam nessa rede.

    Não depende das cargas: o inversor on-grid injeta na rede, não alimenta
    carga em ilha. Usa o mesmo CONEXOES_REDE que _pick_string_inverter usa.
    """
    conexoes = CONEXOES_REDE.get(padrao, {})
    atendem = []
    for inv in inversores:
        fase = str(eff(inv, "phase") or "")
        tensao = str(eff(inv, "voltage") or "")
        esperadas = conexoes.get(fase)
        if esperadas and (not tensao or tensao in esperadas):
            atendem.append(_tit(inv))
    return atendem


def _resumir(titulos: list[str]) -> str:
    """Lista completa, sem truncar.

    A tabela existe para o engenheiro conferir item a item; um "…e mais 72"
    esconderia justamente o que ele precisa auditar. Célula do Excel comporta
    32 mil caracteres — o pior caso aqui não chega perto.
    """
    if not titulos:
        return "nenhum"
    ordenados = sorted(set(titulos))
    return f"({len(ordenados)})\n" + ";\n".join(ordenados) + ";"


async def montar(marca: str | None = None) -> list[dict]:
    async with AsyncSessionLocal() as db:
        hibridos = await list_products(db, tipo="inversor_hibrido",
                                       marca=marca, active=True)
        strings = await list_products(db, tipo="inversor_string",
                                      marca=marca, active=True)

    # Sem preço não é produto cotável; poluiria a matriz com item de catálogo
    # morto. O critério é o mesmo do motor — preço derivado do CUSTO — senão a
    # matriz prometeria equipamento que a cotação recusa.
    hibridos = [h for h in hibridos if preco_venda(h)]
    strings = [s for s in strings if preco_venda(s)]
    print(f"catálogo: {len(hibridos)} híbridos, {len(strings)} on-grid", file=sys.stderr)

    linhas = []
    for padrao, tensao_rede, ligacao_rede in PADROES:
        ongrid = _ongrid_que_atendem(strings, padrao)
        for tensao_carga, fase_carga, rotulo_fase in CARGAS:
            fase_rede = _NOME_FASES[FASES_REDE[padrao]].replace("á", "a")
            # Cenário que a instalação não comporta não tem "inversor que
            # atende" — tem um erro de premissa. Listar equipamento aqui foi o
            # que fez a primeira versão da matriz parecer que o motor aceitava
            # carga 380 V num padrão de 127 V.
            impossivel = carga_existe_na_rede(padrao, tensao_carga, fase_carga)
            if impossivel:
                hib, ressalva, recusados = [], [], []
            else:
                hib, ressalva, recusados = _hibridos_que_atendem(
                    hibridos, padrao, fase_rede, tensao_carga, fase_carga)
            linhas.append({
                "impossivel": impossivel,
                "tensao_rede": tensao_rede,
                "ligacao_rede": ligacao_rede,
                "tensao_carga": tensao_carga or "—",
                "fase_carga": rotulo_fase,
                "hibridos": hib,
                "hibridos_ressalva": ressalva,
                # On-grid só entra na linha sem cargas: um kit com backup precisa
                # do híbrido, e o string não alimenta carga em ilha.
                "ongrid": ongrid if tensao_carga is None else [],
                "recusados": recusados,
            })
    return linhas


def gravar_xlsx(linhas: list[dict], caminho: str) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compatibilidade"

    cabecalhos = [
        "Tensão do padrão de entrada",
        "Tipo de ligação do padrão de entrada",
        "Tensão das cargas de backup",
        "Tipo de ligação das cargas de backup",
        "Inversores híbridos que atendem",
        "Híbridos que atendem SÓ com mudança do padrão de entrada",
        "Inversores on-grid que atendem",
        "Por que os demais híbridos não atendem",
    ]
    ws.append(cabecalhos)

    negrito = Font(bold=True, color="FFFFFF")
    fundo = PatternFill("solid", fgColor="1F4E78")
    for c in ws[1]:
        c.font = negrito
        c.fill = fundo
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    zebra = PatternFill("solid", fgColor="F2F7FB")
    sem_solucao = PatternFill("solid", fgColor="FCE4E4")

    for i, ln in enumerate(linhas):
        motivos = {}
        for titulo, motivo in ln["recusados"]:
            motivos.setdefault(motivo, []).append(titulo)
        resumo = "\n".join(
            f"• {motivo} ({len(ts)}): {', '.join(sorted(ts)[:3])}"
            + ("…" if len(ts) > 3 else "")
            for motivo, ts in sorted(motivos.items(), key=lambda kv: -len(kv[1]))
        ) or "—"

        if ln["impossivel"]:
            ws.append([
                ln["tensao_rede"], ln["ligacao_rede"],
                ln["tensao_carga"], ln["fase_carga"],
                "CENÁRIO IMPOSSÍVEL", "—", "—",
                f"A instalação não comporta essa carga: {ln['impossivel']}. "
                f"O motor recusa a cotação antes de escolher equipamento.",
            ])
        else:
            ws.append([
                ln["tensao_rede"], ln["ligacao_rede"],
                ln["tensao_carga"], ln["fase_carga"],
                _resumir(ln["hibridos"]),
                _resumir(ln["hibridos_ressalva"]) if ln["hibridos_ressalva"] else "—",
                _resumir(ln["ongrid"]) if ln["ongrid"] else "—",
                resumo,
            ])
        linha_xl = i + 2
        vazia = bool(ln["impossivel"])
        for c in ws[linha_xl]:
            c.alignment = Alignment(vertical="top", wrap_text=True)
            if vazia:
                c.fill = sem_solucao
            elif i % 2:
                c.fill = zebra

    for col, larg in zip("ABCDEFGH", (16, 22, 16, 22, 46, 40, 46, 56)):
        ws.column_dimensions[col].width = larg
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cabecalhos))}{ws.max_row}"

    _aba_regras(wb)
    wb.save(caminho)


def _aba_regras(wb) -> None:
    """Aba com a regra por trás de cada coluna, e de onde ela vem no código.

    Sem isso a matriz é uma lista de nomes: o engenheiro consegue discordar,
    mas não consegue apontar ONDE discorda.
    """
    from openpyxl.styles import Alignment, Font

    ws = wb.create_sheet("Regras")
    linhas = [
        ("Como esta tabela foi gerada", ""),
        ("", "Cada linha chama as funções do próprio motor de cálculo "
             "(backend/app/engines/kit_builder.py). Nenhuma regra foi "
             "reimplementada aqui — se o motor mudar, a tabela muda junto. "
             "Regenerar com: python scripts/matriz_compatibilidade.py --xlsx <arquivo>"),
        ("", ""),
        ("O que NÃO está considerado", ""),
        ("Potência", "A matriz é de compatibilidade ELÉTRICA, não de "
                     "dimensionamento. Não olha kW nem kWh: qualquer modelo "
                     "compatível pode ser escalado em paralelo (R4) ou receber "
                     "mais baterias (R1/R2). Quem escolhe o modelo e a "
                     "quantidade é a tabela de cenários."),
        ("Preço", "Idem — a escolha por preço só acontece entre os que já "
                  "passaram por estes filtros."),
        ("", ""),
        ("Regra aplicada em cada coluna", ""),
        ("Inversores híbridos que atendem",
         "Passa em _bloqueio_rede (R7: inversor trifásico precisa operar na "
         "tensão entre fases da rede) E em compativel_com_cargas (R8: tensão e "
         "fase das cargas), sem alerta de rede pendente."),
        ("Híbridos que atendem SÓ com mudança do padrão de entrada",
         "Compatíveis com a carga, mas trifásicos numa unidade monofásica. O "
         "motor NÃO bloqueia — emite alerta. Exige aumento de carga / troca do "
         "padrão junto à concessionária antes de vender."),
        ("Inversores on-grid que atendem",
         "Usa CONEXOES_REDE (backend/app/calculate/service.py). Não depende das "
         "cargas: o on-grid injeta na rede, não alimenta carga em ilha — por "
         "isso só a linha 'sem cargas de backup' traz on-grid."),
        ("Por que os demais híbridos não atendem",
         "Motivo devolvido pelo próprio motor, agrupado. É o mesmo texto que "
         "aparece no painel 'Por que este kit' da ferramenta."),
        ("", ""),
        ("Conceitos que decidem as linhas", ""),
        ("Tensão entre fases × tensão de fase",
         "'380/220' quer dizer 380 V entre fases e 220 V entre fase e neutro. "
         "Carga TRIFÁSICA precisa da tensão entre fases (ocupa as três, não dá "
         "para remanejar). Carga BIFÁSICA usa dois condutores e só precisa da "
         "diferença de potencial correta — num 380/220 ela é ligada entre fase "
         "e neutro e recebe 220 V."),
        ("Saída selecionável das linhas K",
         "O K008/K017 é cadastrado '380/220;220/127': duas configurações, mas "
         "UMA seleção. É o mesmo enrolamento que se liga na rede e alimenta o "
         "EPS. Fixado o padrão de entrada, sobra uma configuração — por isso o "
         "K serve carga trifásica 220 V numa rede 127/220 e carga trifásica "
         "380 V numa rede 220/380, nunca as duas na mesma instalação."),
        ("Monofásico em rede trifásica",
         "Conexão válida, fase-neutro, e costuma ser mais barata que um "
         "trifásico equivalente. Com nº de unidades não múltiplo de 3 a geração "
         "fica desequilibrada entre as fases — o motor emite alerta."),
    ]
    for titulo, texto in linhas:
        ws.append([titulo, texto])
    for c in ws["A"]:
        c.font = Font(bold=True)
        c.alignment = Alignment(vertical="top", wrap_text=True)
    for c in ws["B"]:
        c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 96


def main() -> None:
    marca = None
    if "--marca" in sys.argv:
        marca = sys.argv[sys.argv.index("--marca") + 1]
    linhas = asyncio.run(montar(marca))

    if "--xlsx" in sys.argv:
        destino = sys.argv[sys.argv.index("--xlsx") + 1]
        gravar_xlsx(linhas, destino)
        print(f"planilha gravada: {destino}")
        return

    for ln in linhas:
        print(f"\n{'=' * 70}\nrede {ln['tensao_rede']} {ln['ligacao_rede']} | "
              f"carga {ln['tensao_carga']} {ln['fase_carga']}")
        print(f"  ressalva ({len(ln['hibridos_ressalva'])})")
        print(f"  híbridos ({len(ln['hibridos'])}): "
              f"{', '.join(sorted(ln['hibridos'])) or 'nenhum'}")
        if ln["ongrid"]:
            print(f"  on-grid  ({len(ln['ongrid'])})")


if __name__ == "__main__":
    if not os.getenv("DATABASE_URL"):
        sys.exit("defina DATABASE_URL")
    main()
